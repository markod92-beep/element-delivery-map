#!/usr/bin/env python3
"""
process_payworks.py
-------------------
Builds driver_hours.json for the Daily Truck Tracker: a per-driver, per-date
lookup of ACTUAL hours worked, sourced from the Payworks timesheet export.

The tracker matches each truck's dispatch driver name to a Payworks employee
and auto-fills Actual Lab Hrs for the selected day from this file.

Source:  Element Event Solutions - Timesheet.xlsx  (sheet "2026 YTD ...")
Output:  driver_hours.json  (deployed alongside the map)

Hours per driver/date = sum of all earning rows (Regular + Overtime),
deduped by Timesheet_Id so accidental duplicate rows are not double-counted.
Only employees who hold a Driver / Driver Helper occupation are included.
"""
import openpyxl, json, datetime, sys, glob, os, re

SRC = os.environ.get("PAYWORKS_XLSX") or (
    glob.glob("/sessions/great-eager-goodall/mnt/Payworks/*Timesheet*.xlsx") or
    glob.glob(r"C:\Users\MarkO'Donoghue\OneDrive - Element Events\Data Platform File Storage - Daily Data Extracts\Payworks\*Timesheet*.xlsx")
)[0]
OUT = os.environ.get("DRIVER_HOURS_OUT", os.path.join(os.path.dirname(os.path.abspath(__file__)), "driver_hours.json"))

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
# pick the most recent YTD sheet (e.g. "2026 YTD (Jun 23)")
ytd = [s for s in wb.sheetnames if "YTD" in s.upper()]
sheet_name = ytd[0] if ytd else wb.sheetnames[0]
ws = wb[sheet_name]

# column indices (0-based) from the export header
C_TSID, C_EMPID, C_FIRST, C_LAST, C_OCC, C_EARN, C_START, C_HOURS = 0, 2, 3, 4, 6, 7, 8, 13

emps = {}            # emp_id -> {first,last,occ:set, dates:{date:{tsid:hours}}}
for i, r in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    eid = r[C_EMPID]
    if eid is None or r[C_FIRST] is None:
        continue
    try:
        hrs = float(r[C_HOURS])
    except (TypeError, ValueError):
        continue
    sd = r[C_START]
    if not isinstance(sd, datetime.datetime):
        continue
    d = sd.date().isoformat()
    e = emps.setdefault(eid, {"first": str(r[C_FIRST]), "last": str(r[C_LAST] or ""),
                              "occ": set(), "dates": {}})
    if r[C_OCC]:
        e["occ"].add(str(r[C_OCC]))
    # dedupe by Timesheet_Id within a date, then sum across earning rows
    e["dates"].setdefault(d, {})[r[C_TSID]] = hrs

def is_driver(occ):
    return any("DRIVER" in o.upper() for o in occ)

drivers = []
for eid, e in emps.items():
    if not is_driver(e["occ"]):
        continue
    by_date = {d: round(sum(tsmap.values()), 2) for d, tsmap in e["dates"].items()}
    drivers.append({
        "f": e["first"].strip().upper(),
        "l": e["last"].strip().upper(),
        "occ": sorted(e["occ"])[0] if e["occ"] else "",
        "d": by_date,
    })

drivers.sort(key=lambda x: (x["f"], x["l"]))
out = {
    "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
    "source_sheet": sheet_name,
    "driver_count": len(drivers),
    "drivers": drivers,
}
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(out, f, separators=(",", ":"))

print(f"Source sheet: {sheet_name}")
print(f"Driver employees written: {len(drivers)}")
print(f"Output: {OUT}  ({os.path.getsize(OUT)/1024:.0f} KB)")
