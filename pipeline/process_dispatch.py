#!/usr/bin/env python3
"""process_dispatch.py — convert CMMC_Dispatch_Routes.xlsx → dispatch_data.json

Two-sheet xlsx from POR:
  Map_Route            (one row per truck route on a given day)
  Map_Route_Details    (one row per stop, joined to route via Route_Number)

Output JSON shape (consumed by index.html Routes layer):
{
  "generated_at": "2026-04-17T10:00:00",
  "window": {"start": "2026-01-17", "end": "2026-05-17"},
  "dates": ["2026-01-17", ...],                 # sorted list of dates that have routes
  "routes": {
    "<date>": [
      {
        "rn":    <route_number>,
        "truck": "ERG-955",
        "drv":   "Bob Smith",
        "stps":  <stop count>,
        "dist":  <km>,
        "time":  <minutes>,
        "load":  <cubes>,
        "hs":    "002",           # home store ('000' placeholders now allowed through)
        "start": [lat, lng],       # depot start coord (if sane)
        "stop":  [lat, lng],       # depot end coord (if sane)
        "geom":  "<enc_polyline>", # optional: OSRM road-following polyline (precision 5)
        "color": "#RRGGBB",        # stable truck color (hash of truck name)
        "stops": [
          {
            "seq":  <route sequence>,
            "tt":   "Delivery|Pick Up|Sent Out|Store Transfer|...",
            "cls":  "D|P|S|T",      # Delivery | Pickup | Subrent | Transfer
            "con":  "<contract>",
            "cust": "<customer>",
            "store":"<CMM|ERG|...>",
            "rev":   <net_revenue>,
            "ll":   [lat, lng],
            "ot":   "<OnTime_Status>",
            "parr": "HH:MM",        # Promised_Arrival
            "aarr": "HH:MM",        # Actual_Arrival
            "addr": "<short address>"
          }, ...
        ]
      }, ...
    ]
  }
}

Keeping JSON compact because dispatch data spans 120 days.

Filters (updated 2026-07-06 — see PART 1 of CLAUDE_CODE_FILTER_FIX):
  - status in {'AX','CX','PX','FX'} → excluded (cancelled)
  - Trip_Date outside [today-90, today+30] → excluded
  - Document_Type == 'Transfer' stops → excluded (internal store transfers)
  - Internal legs by customer name (SCHEDULED STOP / STOP AT STORE / STR #0.. /
    '... ELEMENT EVENT SOLUTIONS ...') → excluded
  - Trip class 'T' (Store-Transfer legs) → excluded (belt-and-suspenders)
  - Route must have ≥1 geocoded stop (lat,lng both non-zero)
  - Home_Store == '000' routes are NO LONGER blanket-dropped: they carry real
    customer stops parented under placeholder headers (e.g. route 55968 /
    contract L12828, Amanda Jerome Events). Genuinely-empty '000' placeholder
    rows still fall out via the zero-geocoded-stop rule at end of process().
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
import sys
import time
import urllib.error
import urllib.request
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

# pypostalcode — Canadian FSA centroid DB. Used to sanity-check POR-supplied
# lat/lng against the delivery postal code. Catches cases like contract
# 689583 (2026-04-20) where POR wrote Detroit coords onto an Oakville (L6K)
# delivery because the user clicked an unverified Google suggestion.
try:
    from pypostalcode import PostalCodeDatabase
    _PC_DB = PostalCodeDatabase()
except Exception as _pc_err:  # pragma: no cover
    # DO NOT let this fail silently. When pypostalcode is missing, BOTH the
    # coord sanity-check and the FSA-centroid fallback become no-ops: bad POR
    # coords sail through (see the 172 stops parked on the Ontario province
    # centroid, 2026-07-09) and coord-less stops get dropped from the map.
    # It was omitted from requirements.txt during the cloud migration and this
    # try/except hid it for days. Shout about it.
    _PC_DB = None
    print("CRITICAL: pypostalcode unavailable (%s) -- geocode validation and "
          "FSA fallback are DISABLED. Add it to pipeline/requirements.txt."
          % _pc_err, file=sys.stderr)
    print("::error title=Geocode guard disabled::pypostalcode missing; "
          "bad coordinates will not be corrected", file=sys.stderr)


# ============================================================
# Venue overrides — shared loader from venue_utils.py (replaces the
# copy-pasted _VENUE_OVERRIDES / _load_venue_overrides pattern).
# ============================================================
from venue_utils import load_venue_overrides, match_venue_override_for_stop


# POR writes the ONTARIO PROVINCE CENTROID when its own geocoder fails, rather
# than leaving the coord blank. Left alone it parks stops ~864 km north of their
# real address (2026-07-09: 172 stops, incl. contract 704636 "550 Bayview Ave,
# Toronto"). Treat it as "no coordinate" so the venue-override / FSA-centroid
# fallback chain can rescue the stop.
_SENTINEL_COORDS = (
    (50.074657, -85.828735),   # centroid of Ontario
)


def _is_sentinel_coord(lat: float, lng: float, tol: float = 0.01) -> bool:
    """True if (lat,lng) is one of POR's 'could not geocode' placeholder coords.
    tol=0.01 deg is roughly 1 km, tight enough to never match a real address."""
    return any(abs(lat - a) < tol and abs(lng - b) < tol for a, b in _SENTINEL_COORDS)


def _extract_fsa(zip_str: str) -> str | None:
    """Return the 3-char Canadian FSA from a postal-code string, or None.
    FSA = letter-digit-letter (e.g. 'L6K 3B3' -> 'L6K')."""
    if not zip_str:
        return None
    s = zip_str.strip().upper().replace(" ", "")
    if len(s) < 3:
        return None
    head = s[:3]
    if head[0].isalpha() and head[1].isdigit() and head[2].isalpha():
        return head
    return None


def _fsa_from_zip_or_selected(zip_str: str, selected_addr: str) -> str | None:
    """FSA from Delivery_Zip, falling back to a Canadian postal parsed out of
    Selected_Address. POR sometimes leaves Delivery_Zip blank on a real stop
    while the picked Selected_Address still carries the postal (route 55968 /
    L12828 Amanda Jerome Events: zip blank, Selected_Address '... N0H 1J0,
    Canada'). Used only as a last-ditch geocode when POR gave no lat/lng and no
    venue override matched."""
    fsa = _extract_fsa(zip_str)
    if fsa:
        return fsa
    if selected_addr:
        m = _CA_POSTAL_RE.search(selected_addr.upper())
        if m:
            return _extract_fsa(m.group(0))
    return None


def _haversine_km(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    R = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lng2 - lng1)
    a = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return 2 * R * math.asin(math.sqrt(a))


_CA_POSTAL_RE = re.compile(r"\b[A-Z]\d[A-Z]\s*\d[A-Z]\d\b")
_US_MARKERS = ("UNITED STATES", " USA", ", USA")


def _selected_address_is_non_canadian(selected_addr: str) -> bool:
    """Return True only when we're confident Selected_Address points to a
    non-Canadian location. False means 'looks Canadian' OR 'unknown' — we
    must not override POR coord on unknowns.

    Strong positive signals:
      - Contains 'United States' / 'USA'
      - Missing a Canadian postal-code pattern (e.g. 'V8E 0Z5')
      - Has neither 'Canada' nor a Canadian postal — ambiguous, skip override
    """
    if not selected_addr:
        return False   # no signal, don't override
    s = selected_addr.upper()
    if any(m in s for m in _US_MARKERS):
        return True
    # Confirmed Canadian if it has a CA postal pattern or literal 'CANADA'
    if _CA_POSTAL_RE.search(s) or "CANADA" in s:
        return False
    # Unknown — be conservative, don't override
    return False


def _validate_against_fsa(lat: float, lng: float, zip_str: str,
                          selected_addr: str) -> tuple[float, float, str]:
    """Sanity-check POR lat/lng against the contract's Delivery_Zip.

    Override POR coord only when ALL three conditions hold:
      1. Delivery_Zip parses to a known Canadian FSA
      2. POR coord is >FSA_MISMATCH_THRESHOLD_KM from that FSA centroid
      3. Selected_Address is clearly non-Canadian (contains 'United States'
         or lacks a Canadian postal-code pattern / 'Canada' marker)

    Condition #3 prevents false positives on rural FSAs (e.g. V0N / Whistler)
    where the centroid sits far from the actual town but the POR coord is
    correct. Root-case: contract 689583 (2026-04-20) — Oakville L6K address
    landed in Detroit because POR's address-picker returned a US address
    ('518 W Service Dr, Detroit, MI, 48242, United States') that the user
    never verified.

    Returns (lat, lng, source) where source is 'por' (POR coord accepted,
    default) or 'fsa_fallback' (snapped to FSA centroid).
    """
    if _PC_DB is None:
        return lat, lng, "por"
    fsa = _extract_fsa(zip_str)
    if not fsa:
        return lat, lng, "por"
    try:
        info = _PC_DB[fsa]
    except Exception:
        return lat, lng, "por"
    c_lat = float(info.latitude)
    c_lng = float(info.longitude)
    dist = _haversine_km(lat, lng, c_lat, c_lng)
    if dist <= FSA_MISMATCH_THRESHOLD_KM:
        return lat, lng, "por"
    # Distance mismatch — require a second signal before overriding
    if _selected_address_is_non_canadian(selected_addr):
        return c_lat, c_lng, "fsa_fallback"
    return lat, lng, "por"


def truck_color(name: str) -> str:
    if not name:
        return "#6B7A7E"
    h = hashlib.md5(name.encode("utf-8")).digest()
    return TRUCK_COLORS[h[0] % len(TRUCK_COLORS)]


def to_date(val) -> date | None:
    """Parse a cell value as a date. Drops sentinel years (<2000) that POR emits
    when a user enters a time without a date (Excel stores those as 1899-12-30)."""
    d = None
    if val is None:
        return None
    if isinstance(val, datetime):
        d = val.date()
    elif isinstance(val, date):
        d = val
    else:
        return None
    if d and d.year < 2000:
        return None
    return d


def fmt_hhmm(val) -> str:
    """datetime/time/string → 'HH:MM', blank otherwise.

    POR's Dispatch export stores Promised/Actual arrival cells as *strings*
    in the form 'YYYY-MM-DD HH:MM' (not datetime), with '1899-12-30 ...' as
    the sentinel for "no value". We drop the sentinel and extract HH:MM.
    """
    if val is None:
        return ""
    if isinstance(val, datetime):
        if val.year < 2000:   # Excel epoch sentinel = "no value"
            return ""
        return val.strftime("%H:%M")
    s = str(val).strip()
    if not s:
        return ""
    # Drop POR's "no value" sentinel (Excel epoch for time-only entries)
    if s.startswith("1899-"):
        return ""
    # 'YYYY-MM-DD HH:MM[:SS]' → take the HH:MM after the space
    if " " in s and ":" in s:
        tail = s.split(" ", 1)[1]
        if len(tail) >= 4 and ":" in tail:
            return tail[:5]
    # Pure 'HH:MM[:SS]' or 'H:MM' time string
    if ":" in s and len(s) >= 4 and s[0].isdigit():
        return s[:5]
    return ""


def num(val, default=0.0) -> float:
    try:
        if val is None:
            return default
        return float(val)
    except (TypeError, ValueError):
        return default


def non_blank(val) -> str:
    return str(val).strip() if val is not None else ""


# ============================================================
# Truck-name parsing → (tno, ttype). Emitted per route alongside the
# existing "truck"/"tnum" fields so the tracker UI can show a number +
# type/size pill (e.g. "#834 · Ryder 5-ton"). The raw display truck name
# is split into a number ("tno") and a type/size label ("ttype").
# Synthetic / non-truck buckets pass through with both blank.
# ============================================================

# Optional per-fleet-prefix size labels. CMM / ERG / DF truck sizes go here
# later once the fleet roster is confirmed (e.g. {"CMM": "5-ton"}); leave
# empty for now so ttype is just the bare fleet code (e.g. "CMM").
FLEET_SIZE: dict = {}

# BC bracket-size code → human label.
_BC_SIZE = {
    "3T": "3-ton", "4T": "4-ton", "5T": "5-ton",
    "4TLO": "4-ton", "VAN": "Van",
}

# Synthetic / non-truck route labels: no parseable number or type.
_SYNTHETIC_TRUCKS = {
    "Unknown", "Large Runs", "Small Runs", "Transfers",
    "Product for Internal use", "",
}

# Truck buckets excluded entirely from the tracker (screen + data). These are
# non-delivery logistics / store-transfer runs, not customer deliveries.
# Mark's call 2026-07-06: KEEP "Large Runs"/"Small Runs" — those carry ~$845K of
# real customer deliveries/pickups that just have no truck # assigned — but drop
# "Transfers" and "Product for Internal use". Individual Store-Transfer legs
# (trip class "T") are also stripped from every route (see Pass 2).
EXCLUDE_TRUCKS = {"Transfers", "Product for Internal use"}


# Internal / non-customer stop legs, matched on Customer_Name. These ride on
# real routes (any Document_Type / Trip_Type) but are not customer deliveries:
#   - "SCHEDULED STOP"            — driver's internal scheduling placeholder
#   - "STOP AT STORE 005 - TRAPP" — internal store visit
#   - "STR #001 Element Event Solutions ..." — store-to-store transfer origin
#   - "... Element Event Solutions ..."      — any Element store name
# Added 2026-07-06 (PART 1 filter fix). Precise so real customers on Logistic
# docs (Royal Ontario Museum, The Food Dudes, Amanda Jerome Events, ...) stay.
_INTERNAL_STORE_RE = re.compile(r"\bSTR\s*#?\s*0\d")


def is_internal_stop(customer_name: str) -> bool:
    """True when a stop's Customer_Name marks it as an internal (non-customer)
    leg — a scheduled placeholder, a store visit, or an Element store transfer."""
    c = (customer_name or "").upper()
    if not c:
        return False
    if "SCHEDULED STOP" in c:
        return True
    if "STOP AT STORE" in c:
        return True
    if "ELEMENT EVENT SOLUTIONS" in c:
        return True
    if _INTERNAL_STORE_RE.search(c):
        return True
    return False


def _trailing_digits(s: str) -> str:
    """Trailing run of digits in s, or '' if none ('BC-RENT1' -> '1')."""
    m = re.search(r"(\d+)\s*$", s)
    return m.group(1) if m else ""


def parse_truck(truck_name: str) -> tuple[str, str]:
    """Parse a (display) truck name into (tno, ttype).

    tno   = truck number as a string ('' when none)
    ttype = type/size label   ('' when none)

    Examples:
        "Ryder 5ton 834"  -> ("834", "Ryder 5-ton")
        "Ryder Cube 179"  -> ("179", "Ryder Cube")
        "Sprinter Van 1"  -> ("1",   "Sprinter Van")
        "BC-05 [4T]"      -> ("05",  "BC 4-ton")
        "BC-FIFI"         -> ("",    "BC-FIFI")
        "CMM-896"         -> ("896", "CMM")
        "#175"            -> ("175", "")
        "Unknown"         -> ("",    "")
    """
    name = (truck_name or "").strip()

    # Synthetic / non-trucks — no number, no type.
    if name in _SYNTHETIC_TRUCKS or name.startswith("No truck assigned"):
        return "", ""

    # Bare number, optionally '#'-prefixed: "#175" / "175".
    m = re.fullmatch(r"#?\s*(\d+)", name)
    if m:
        return m.group(1), ""

    # Ryder: "Ryder 5ton 834", "Ryder Cube 179", "Ryder Ford Van 995".
    m = re.fullmatch(r"Ryder\s+(.*?)\s+(\d+)", name)
    if m:
        words = m.group(1).strip()
        # Normalize "5ton" / "5-ton" → "5-ton".
        words = re.sub(r"\b5-?ton\b", "5-ton", words, flags=re.IGNORECASE)
        return m.group(2), f"Ryder {words}"

    # Sprinter Van: "Sprinter Van 1".
    m = re.fullmatch(r"Sprinter\s+Van\s+(\d+)", name, flags=re.IGNORECASE)
    if m:
        return m.group(1), "Sprinter Van"

    # BC with bracket size: "BC-05 [4T]".
    m = re.fullmatch(r"BC-(\w+)\s*\[(\w+)\]", name)
    if m:
        size = _BC_SIZE.get(m.group(2).upper(), m.group(2))
        return m.group(1), f"BC {size}"

    # BC without a bracket: "BC-RENT1", "BC-FIFI".
    if name.startswith("BC-"):
        return _trailing_digits(name), name

    # Fleet code + number: "CMM-896", "ERG-954", "DF-1362", "BR66434", "van-1".
    m = re.fullmatch(r"([A-Za-z]+)-?(\d+)", name)
    if m:
        prefix = m.group(1)
        ttype = prefix
        size = FLEET_SIZE.get(prefix.upper())
        if size:
            ttype = f"{prefix} {size}"
        return m.group(2), ttype

    # Anything else: trailing digits if present, name as-is for type.
    return _trailing_digits(name), name


# ============================================================
# Module constants — restored 2026-04-29 from the Apr 21
# __pycache__/process_dispatch.cpython-310.pyc after an Apr 24
# refactor stripped them. Values are verbatim from the last
# working compiled bytecode (extracted via xdis from co_consts /
# module-level disassembly). Do not edit without verifying against
# the .pyc unless you know what you're changing.
# ============================================================

# Map Trip_Type → single-letter class used by the UI
# (D = Delivery, P = Pickup, S = Subrent, T = Transfer).
TRIP_CLASS = {
    "Delivery":             "D",
    "L-Delivery":           "D",
    "L-Delv Done":          "D",
    "Pick Up":              "P",
    "L-Pick Up":            "P",
    "L-Pick Done":          "P",
    "Returned":             "P",
    "Sent Out":             "S",
    "Pickup Subrent":       "S",
    "Return Subrent":       "S",
    "Subrent In":           "S",
    "Subrent Returned":     "S",
    "Store Transfer":       "T",
    "Transferred To Store": "T",
    "New":                  "D",
    "Repair Done":          "T",
}

# Cancelled-route status codes — routes with these in Map_Route.Status
# are dropped at ingest (matches POR's cancelled/voided semantics).
DROP_ROUTE_STATUS = {"AX", "CX", "PX", "FX"}

# Stable per-truck palette. truck_color() hashes the truck name
# and indexes into this list, so each truck keeps the same color
# across daily runs without needing a persisted mapping.
TRUCK_COLORS = [
    "#0B4F6C", "#D94E1F", "#1B7A3A", "#7A3E9D",
    "#C9A227", "#2E86AB", "#A23B72", "#17A398",
    "#F18F01", "#5A5B9F", "#8B5A2B", "#6B7A7E",
]

# Depot lat/lng per home-store code. Used as the start/stop fallback
# when POR doesn't supply usable StartStore_*/StopStore_* coords on a
# Map_Route row, so polylines anchor at the warehouse instead of the
# first customer stop. Sanity-checked against the route's stop centroid
# (THRESH_DEG ≈ 400 km) before being applied — prevents pasting a
# Toronto depot onto a Halifax route.
HOME_STORE_COORDS = {
    "001": (43.7783, -79.3374),
    "002": (43.7063, -79.3603),
    "003": (44.6852, -63.4959),
    "004": (43.7063, -79.3603),
    "005": (49.2183, -122.9594),
    "006": (43.2025, -79.7353),
    "007": (49.2770, -123.0755),
    "008": (43.7063, -79.3603),
    "009": (43.6263, -79.5328),
}

# Distance (km) above which a POR-supplied lat/lng is considered to
# disagree with the contract's Delivery_Zip FSA centroid. Triggers a
# fallback to the FSA centroid only when Selected_Address also looks
# non-Canadian. Root-case: contract 689583 (2026-04-20), Oakville L6K
# address with POR coords pointing to Detroit.
FSA_MISMATCH_THRESHOLD_KM = 100.0


# --- OSRM driving geometry (road-following polylines) ---
#
# We call the public OSRM demo server by default (`router.project-osrm.org`) to
# convert a sequence of [lat, lng] waypoints into an encoded-polyline string
# that follows real roads. Results are cached to disk (`osrm_cache.json`, next
# to dispatch_data.json) keyed on the rounded waypoint sequence, so a given
# route is fetched at most once even across daily refreshes.
#
# Design notes:
#   - Cache-hit is a pure local lookup → free and offline-friendly.
#   - Cache-miss sleeps ~0.25s between calls to be polite to the demo server
#     (Mark's machine runs this overnight, speed isn't critical).
#   - Any fetch failure returns None; the caller falls back to straight lines.
#     Missing geometry never blocks the pipeline.
#   - Encoded polyline (precision 5) is ~1/10th the size of raw lat/lng JSON,
#     so shipping it to the browser is cheap.
#   - `r.geom` on the output route object = encoded polyline string, or absent
#     if the fetch never succeeded.

OSRM_CACHE_VERSION = 1
OSRM_MAX_WAYPOINTS = 100   # OSRM demo limit is ~100; over that we skip.
OSRM_HTTP_TIMEOUT  = 8.0   # seconds per request
OSRM_SLEEP_BETWEEN = 0.25  # seconds between live fetches (cache hits = 0)


def _osrm_cache_key(waypoints) -> str:
    """Deterministic MD5 hash of waypoint list (rounded to 5 decimal places)."""
    rounded = ";".join(f"{lat:.5f},{lng:.5f}" for lat, lng in waypoints)
    return hashlib.md5(rounded.encode("utf-8")).hexdigest()


def _osrm_load_cache(path: Path) -> dict:
    if not path.exists():
        return {"v": OSRM_CACHE_VERSION, "entries": {}}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        if data.get("v") != OSRM_CACHE_VERSION or "entries" not in data:
            return {"v": OSRM_CACHE_VERSION, "entries": {}}
        return data
    except Exception:
        return {"v": OSRM_CACHE_VERSION, "entries": {}}


def _osrm_save_cache(path: Path, cache: dict) -> None:
    try:
        path.write_text(json.dumps(cache, separators=(",", ":")), encoding="utf-8")
    except Exception as e:
        print(f"[dispatch] WARN: could not save OSRM cache: {e}", file=sys.stderr)


def _fetch_osrm_geometry(waypoints, osrm_url: str, cache: dict) -> tuple[str | None, bool]:
    """Return (encoded_polyline_or_None, was_cache_hit).

    Failure modes (always return (None, was_cache_hit)):
        - <2 waypoints
        - >OSRM_MAX_WAYPOINTS waypoints
        - HTTP timeout / connection error / non-Ok response
        - Any JSON parse / shape problem

    Negative cache entries are stored too, so we don't re-hammer the server
    for the same dead route on every daily run.
    """
    if len(waypoints) < 2 or len(waypoints) > OSRM_MAX_WAYPOINTS:
        return None, False

    key = _osrm_cache_key(waypoints)
    if key in cache["entries"]:
        return cache["entries"][key].get("geom"), True

    # OSRM expects lng,lat (GeoJSON order) — swap from our lat,lng.
    coords = ";".join(f"{lng:.5f},{lat:.5f}" for lat, lng in waypoints)
    url = (f"{osrm_url.rstrip('/')}/route/v1/driving/{coords}"
           f"?overview=full&geometries=polyline")

    try:
        req = urllib.request.Request(url, headers={"User-Agent": "element-delivery-map/1.0"})
        with urllib.request.urlopen(req, timeout=OSRM_HTTP_TIMEOUT) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except (urllib.error.URLError, TimeoutError, ValueError) as e:
        # Don't cache transient network errors — next run can retry.
        return None, False
    except Exception as e:
        return None, False

    if data.get("code") != "Ok":
        # Cache negative result so we don't retry dead routes every day.
        cache["entries"][key] = {"geom": None, "err": data.get("code")}
        return None, False

    routes = data.get("routes") or []
    if not routes:
        cache["entries"][key] = {"geom": None, "err": "no_routes"}
        return None, False

    geom = routes[0].get("geometry")
    cache["entries"][key] = {"geom": geom}
    return geom, False


# ============================================================
# Map_Route / Map_Route_Details column index resolution.
# Column constants (R_*, D_*) are derived from the actual xlsx headers at
# runtime, so the script survives POR column reorders. If POR renames a
# column, _build_col_map raises with a clear list of actual headers so the
# fix is just adding the new name to the appropriate candidate list.
#
# Restored 2026-04-29: an Apr 24 refactor stripped the previously hardcoded
# R_*/D_* constants without replacing them, breaking every run since
# (NameError on R_ROUTE_NUMBER at process() pass 1). Live site served
# Apr 23 data for 6 days as a result.
# ============================================================

ROUTE_HEADER_SPECS = {
    "ROUTE_NUMBER":  ["Route_Number", "RouteNumber", "Route Number"],
    "HOME_STORE":    ["Home_Store", "HomeStore", "Home Store"],
    "STATUS":        ["Status", "Route_Status", "RouteStatus"],
    "TRIP_DATE":     ["Trip_Date", "TripDate", "Trip Date"],
    "TRUCK_NAME":    ["Truck_Name", "TruckName", "Truck Name"],
    "TRUCK_NUMBER":  ["Truck_Number", "TruckNumber", "Truck Number"],
    "DRIVER_NAME":   ["Driver_Name", "DriverName", "Driver Name", "Operator_Name", "Operator Name"],
    "START_LONG":    ["StartStore_Long", "Start_Longitude", "StartLongitude", "Start_Long", "StartLong", "Starting_Longitude"],
    "START_LAT":     ["StartStore_Lat",  "Start_Latitude",  "StartLatitude",  "Start_Lat",  "StartLat",  "Starting_Latitude"],
    "STOP_LONG":     ["StopStore_Long",  "Stop_Longitude",  "StopLongitude",  "Stop_Long",  "StopLong",  "End_Longitude", "Ending_Longitude"],
    "STOP_LAT":      ["StopStore_Lat",   "Stop_Latitude",   "StopLatitude",   "Stop_Lat",   "StopLat",   "End_Latitude",  "Ending_Latitude"],
    "STOP_COUNT":    ["Stop_Count", "StopCount", "Stop Count", "Stops"],
    "DISTANCE":      ["Route_Distance", "Distance", "Trip_Distance", "TripDistance", "Total_Distance"],
    "TRIP_TIME":     ["Route_Trip_Time", "Trip_Time", "TripTime", "Trip Time", "Total_Time"],
    "LOAD_SIZE":     ["Load_Size", "LoadSize", "Load Size", "Cubes"],
}

DETAIL_HEADER_SPECS = {
    "ROUTE_NUMBER":   ["Route_Number", "RouteNumber", "Route Number"],
    "LATITUDE":       ["Latitude", "Lat"],
    "LONGITUDE":      ["Longitude", "Long", "Lng"],
    "CUSTOMER_NAME":  ["Customer_Name", "CustomerName", "Customer Name"],
    "ADDRESS1":       ["Address1", "Address_1", "Address Line 1", "Delivery_Address", "Address"],
    "CITY":           ["City", "Delivery_City"],
    "DELIVERY_ZIP":   ["Delivery_Zip", "DeliveryZip", "Delivery Zip", "Postal_Code", "PostalCode", "Zip"],
    "SELECTED_ADDR":  ["Selected_Address", "SelectedAddress", "Selected Address"],
    "TRIP_TYPE":      ["Trip_Type", "TripType", "Trip Type"],
    "DOCUMENT_TYPE":  ["Document_Type", "DocumentType", "Document Type"],
    "ROUTE_SEQ":      ["Route_Sequence", "RouteSequence", "Route_Seq", "RouteSeq", "Sequence", "Seq"],
    "CONTRACT":       ["Contract_Number", "ContractNumber", "Contract"],
    "STORE":          ["Store", "Home_Store"],
    "NET_REVENUE":    ["Net_Revenue", "NetRevenue", "Net Revenue"],
    "ONTIME_STATUS":  ["OnTime_Status", "OnTimeStatus", "On_Time_Status", "OnTime Status"],
    "PROMISED_ARR":   ["Promised_Arrival", "PromisedArrival", "Promised Arrival", "Promised_Arr"],
    "ACT_ARRIVAL":    ["Actual_Arrival", "ActualArrival", "Actual Arrival", "Act_Arrival"],
}


def _normalize_header(s) -> str:
    """Case/whitespace/underscore-insensitive header key for fuzzy matching."""
    if s is None:
        return ""
    return str(s).strip().lower().replace(" ", "").replace("_", "").replace("#", "number")


def _build_col_map(ws, sheet_label: str, specs: dict, optional: set | None = None) -> dict:
    """Read row 1 of ws, return {LOGICAL_NAME: column_index}.

    Raises RuntimeError listing the actual headers if any *required* spec can't
    be matched — that error message is the recovery instruction (add the new
    name to the candidate list for that logical column). Logical columns named
    in `optional` are skipped silently when absent (caller uses out.get(...)),
    so a future POR export that drops an optional column degrades gracefully
    instead of hard-failing the live pipeline.
    """
    optional = optional or set()
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise RuntimeError(f"[dispatch] {sheet_label} appears empty (no header row).")
    by_norm = {}
    for i, h in enumerate(header_row):
        n = _normalize_header(h)
        if n and n not in by_norm:  # first occurrence wins on duplicates
            by_norm[n] = i

    out = {}
    missing = []
    for logical, candidates in specs.items():
        idx = None
        for cand in candidates:
            n = _normalize_header(cand)
            if n in by_norm:
                idx = by_norm[n]
                break
        if idx is None:
            if logical in optional:
                continue  # optional column absent — caller handles via .get()
            missing.append(f"  - {logical} (tried: {candidates})")
        else:
            out[logical] = idx
    if missing:
        seen = "\n".join(f"  [{i}] {h!r}" for i, h in enumerate(header_row))
        raise RuntimeError(
            f"[dispatch] {sheet_label}: the following logical columns could not be matched against the header row:\n"
            + "\n".join(missing)
            + "\n\nActual headers found:\n" + seen
            + "\n\nFix: add the actual header name to the candidate list for the missing logical column in process_dispatch.py."
        )
    return out


def process(src_path: str, out_path: str, days_back: int = 90, days_forward: int = 30,
            fetch_geometry: bool = False, osrm_url: str = "https://router.project-osrm.org") -> dict:
    today = date.today()
    window_start = today - timedelta(days=days_back)
    window_end = today + timedelta(days=days_forward)

    print(f"[dispatch] Loading {src_path}...", file=sys.stderr)
    wb = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
    # --- Pass 1: Map_Route -> route metadata (filtered) ---
    route_meta: dict = {}     # route_number -> meta dict
    ws = wb["Map_Route"]
    _R = _build_col_map(ws, "Map_Route", ROUTE_HEADER_SPECS)
    R_ROUTE_NUMBER = _R["ROUTE_NUMBER"]
    R_HOME_STORE   = _R["HOME_STORE"]
    R_STATUS       = _R["STATUS"]
    R_TRIP_DATE    = _R["TRIP_DATE"]
    R_TRUCK_NAME   = _R["TRUCK_NAME"]
    R_TRUCK_NUMBER = _R["TRUCK_NUMBER"]
    R_DRIVER_NAME  = _R["DRIVER_NAME"]
    R_START_LONG   = _R["START_LONG"]
    R_START_LAT    = _R["START_LAT"]
    R_STOP_LONG    = _R["STOP_LONG"]
    R_STOP_LAT     = _R["STOP_LAT"]
    R_STOP_COUNT   = _R["STOP_COUNT"]
    R_DISTANCE     = _R["DISTANCE"]
    R_TRIP_TIME    = _R["TRIP_TIME"]
    R_LOAD_SIZE    = _R["LOAD_SIZE"]
    r_total = r_kept = r_kept_store000 = r_skip_status = r_skip_window = 0
    r_skip_excluded = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        r_total += 1
        rn = row[R_ROUTE_NUMBER]
        if rn is None:
            continue

        # Home_Store == "000" is a placeholder header, NOT a reason to drop the
        # route: real customer stops are parented under these (route 55968 /
        # L12828 Amanda Jerome Events). We let 000 routes through Pass 1 and rely
        # on the end-of-function zero-geocoded-stop rule to remove the genuinely
        # empty placeholders (Stop_Count -1/-2, no real detail stops). (PART 1)
        home = non_blank(row[R_HOME_STORE])
        if home == "000":
            r_kept_store000 += 1

        status = non_blank(row[R_STATUS])
        if status in DROP_ROUTE_STATUS:
            r_skip_status += 1
            continue

        trip_d = to_date(row[R_TRIP_DATE])
        if trip_d is None or trip_d < window_start or trip_d > window_end:
            r_skip_window += 1
            continue

        truck_name = non_blank(row[R_TRUCK_NAME])
        truck_num = non_blank(row[R_TRUCK_NUMBER])
        # Display truck: prefer Truck_Name, fall back to #Truck_Number, then Unknown.
        # Special case: POR users sometimes type a long digit string (rental
        # vendor PO# or truck serial) into the Truck_Name column for rental
        # units that ARE part of Element's fleet (e.g. units 175 / 176 / 177).
        # The tnum column has the clean Element fleet number — prefer that as
        # the pill label so each rental truck keeps its own identity instead
        # of collapsing into a single generic bucket.
        if truck_name:
            if truck_name.isdigit() and len(truck_name) >= 4:
                if truck_num and truck_num != "0":
                    display_truck = f"#{truck_num}"
                else:
                    display_truck = f"#{truck_name}"
            else:
                display_truck = truck_name
        elif truck_num and truck_num != "0":
            display_truck = f"#{truck_num}"
        else:
            display_truck = "Unknown"
        # Drop non-delivery logistics / transfer buckets entirely (screen +
        # data). Excluding here in Pass 1 means their route# never enters
        # route_meta, so Pass 2 stops for these routes are skipped (rn not in
        # route_meta) and they never reach contract_trucks/event_trucks.
        if display_truck in EXCLUDE_TRUCKS:
            r_skip_excluded += 1
            continue

        driver = non_blank(row[R_DRIVER_NAME])
        # Parse the display truck into a number + type/size label for the
        # tracker UI pill. Keeps "truck"/"tnum" unchanged; adds "tno"/"ttype".
        tno, ttype = parse_truck(display_truck)

        # StartStore / StopStore coords from Map_Route — these represent where
        # the truck left from (home warehouse) and where it ended the day. POR
        # sometimes writes zeros here, so we only emit when both lat & lng are
        # non-zero. Downstream sanity-check (vs. stop bbox) happens after pass 2
        # once we actually know where the route's stops are, since a valid depot
        # coord will typically sit within a few dozen km of the stops.
        start_lng = num(row[R_START_LONG])
        start_lat = num(row[R_START_LAT])
        stop_lng  = num(row[R_STOP_LONG])
        stop_lat  = num(row[R_STOP_LAT])

        route_meta[rn] = {
            "rn":    rn,
            "date":  trip_d.isoformat(),
            "truck": display_truck,
            "tnum":  truck_num,
            "tno":   tno,
            "ttype": ttype,
            "drv":   driver,
            "stps":  int(row[R_STOP_COUNT] or 0),
            "dist":  round(num(row[R_DISTANCE]), 1),
            "time":  int(num(row[R_TRIP_TIME])),
            "load":  int(num(row[R_LOAD_SIZE])),
            "hs":    home,
            "color": truck_color(display_truck),
            "stops": [],
            # Raw candidates — may be dropped in post-process if they look like
            # junk (e.g. Burnaby default on a CMM Toronto route).
            "_start_cand": [start_lat, start_lng] if (start_lat and start_lng) else None,
            "_stop_cand":  [stop_lat,  stop_lng]  if (stop_lat  and stop_lng)  else None,
        }
        r_kept += 1

    print(f"[dispatch] Map_Route: total={r_total:,} kept={r_kept:,} "
          f"(incl. store000={r_kept_store000:,} now allowed through) "
          f"skip_cancelled={r_skip_status:,} "
          f"skip_outside_window={r_skip_window:,} "
          f"skip_excluded_truck={r_skip_excluded:,}", file=sys.stderr)

    # --- Pass 2: Map_Route_Details -> attach stops to routes ---
    ws = wb["Map_Route_Details"]
    _D = _build_col_map(ws, "Map_Route_Details", DETAIL_HEADER_SPECS,
                        optional={"DOCUMENT_TYPE"})
    D_ROUTE_NUMBER  = _D["ROUTE_NUMBER"]
    D_LATITUDE      = _D["LATITUDE"]
    D_LONGITUDE     = _D["LONGITUDE"]
    D_CUSTOMER_NAME = _D["CUSTOMER_NAME"]
    D_ADDRESS1      = _D["ADDRESS1"]
    D_CITY          = _D["CITY"]
    D_DELIVERY_ZIP  = _D["DELIVERY_ZIP"]
    D_SELECTED_ADDR = _D["SELECTED_ADDR"]
    D_TRIP_TYPE     = _D["TRIP_TYPE"]
    D_DOCUMENT_TYPE = _D.get("DOCUMENT_TYPE")   # optional — may be None
    D_ROUTE_SEQ     = _D["ROUTE_SEQ"]
    D_CONTRACT      = _D["CONTRACT"]
    D_STORE         = _D["STORE"]
    D_NET_REVENUE   = _D["NET_REVENUE"]
    D_ONTIME_STATUS = _D["ONTIME_STATUS"]
    D_PROMISED_ARR  = _D["PROMISED_ARR"]
    D_ACT_ARRIVAL   = _D["ACT_ARRIVAL"]
    s_total = s_attached = s_no_route = s_no_geo = s_venue_backfill = 0
    s_fsa_fallback = 0
    s_sentinel = 0
    s_skip_transfer = 0
    s_skip_doc_transfer = s_skip_internal = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        s_total += 1
        rn = row[D_ROUTE_NUMBER]
        if rn is None or rn not in route_meta:
            s_no_route += 1
            continue

        lat = num(row[D_LATITUDE])
        lng = num(row[D_LONGITUDE])

        # POR's "couldn't geocode" placeholder -> treat as missing so the
        # venue-override / FSA-centroid fallback below can rescue the stop.
        if lat and lng and _is_sentinel_coord(lat, lng):
            s_sentinel += 1
            lat = lng = 0.0

        # Pre-read the stop identity fields so we can classify the stop and,
        # if it survives, try a venue-override backfill when POR didn't geocode.
        cust_name = non_blank(row[D_CUSTOMER_NAME])
        addr = non_blank(row[D_ADDRESS1])
        city = non_blank(row[D_CITY])
        zip_str = non_blank(row[D_DELIVERY_ZIP])
        selected_addr = non_blank(row[D_SELECTED_ADDR])
        doc_type = non_blank(row[D_DOCUMENT_TYPE]) if D_DOCUMENT_TYPE is not None else ""

        # --- Precise internal-leg exclusion (PART 1) ---
        # Drop internal store transfers and non-customer legs BEFORE any geocode
        # work, so they never render and never anchor a route. Real-customer
        # Contract/Logistic stops (incl. the Amanda Jerome L-Pick Up) fall
        # through and are kept. Transfers are driven primarily by Document_Type;
        # the cls=="T" skip further below stays as belt-and-suspenders.
        if doc_type == "Transfer":
            s_skip_doc_transfer += 1
            continue
        if is_internal_stop(cust_name):
            s_skip_internal += 1
            continue

        geocode_source = "por"
        if not lat or not lng:
            # Try venue-override backfill. Recovers routes like the Apr 15
            # Waterloo L-Delivery (#49866, CMM-867, contract L10976) that
            # would otherwise be silently dropped by the route-must-have-
            # >=1-geocoded-stop filter at the end of this function.
            ov = match_venue_override_for_stop(cust_name, addr, city)
            if ov is not None:
                lat = float(ov["lat"])
                lng = float(ov["lng"])
                geocode_source = "venue_override"
                s_venue_backfill += 1
            else:
                # POR had no coord AND no venue override — try FSA centroid
                # as a last-ditch geocode so the stop still renders. FSA comes
                # from Delivery_Zip, or (when that's blank) from a postal parsed
                # out of Selected_Address — recovers route 55968 / L12828.
                fsa = _fsa_from_zip_or_selected(zip_str, selected_addr)
                if fsa and _PC_DB is not None:
                    try:
                        info = _PC_DB[fsa]
                        lat = float(info.latitude)
                        lng = float(info.longitude)
                        geocode_source = "fsa_fallback"
                        s_fsa_fallback += 1
                    except Exception:
                        s_no_geo += 1
                        continue
                else:
                    s_no_geo += 1
                    continue
        else:
            # POR supplied a coord — sanity-check it against the Delivery_Zip
            # FSA. Catches the 2026-04-20 class of bug (contract 689583:
            # Oakville L6K address, POR lat/lng pointed to Detroit because
            # the user picked an unconfirmed address from a dropdown).
            new_lat, new_lng, src = _validate_against_fsa(lat, lng, zip_str, selected_addr)
            if src == "fsa_fallback":
                lat, lng = new_lat, new_lng
                geocode_source = "fsa_fallback"
                s_fsa_fallback += 1

        trip_type = non_blank(row[D_TRIP_TYPE])
        cls = TRIP_CLASS.get(trip_type, "D")

        # Strip individual Store-Transfer legs (class "T") from every route so no
        # transfer stops render anywhere — even when they ride on a real named
        # truck. Delivery/Pickup/Subrent stops on the same route are unaffected.
        if cls == "T":
            s_skip_transfer += 1
            continue

        short_addr = ", ".join(p for p in [addr, city] if p)

        stop = {
            "seq":   int(row[D_ROUTE_SEQ] or 0),
            "tt":    trip_type,
            "cls":   cls,
            "con":   non_blank(row[D_CONTRACT]),
            "cust":  cust_name,
            "store": non_blank(row[D_STORE]),
            # Map_Route_Details.Net_Revenue == Unique_Events.Net_Revenue (rental,
            # pre-tax, EXCLUDING delivery fee). Verified 2026-04-21 across 90,200
            # dispatch rows joined to UE: 100% match rental, 0 match fee-only.
            # Duplicated across the Delivery + Pickup rows of the same contract,
            # so stop.rev is always the contract rental — never the delivery fee.
            "rev":   round(num(row[D_NET_REVENUE]), 2),
            "ll":    [round(lat, 6), round(lng, 6)],
            "ot":    non_blank(row[D_ONTIME_STATUS]),
            "parr":  fmt_hhmm(row[D_PROMISED_ARR]),
            "aarr":  fmt_hhmm(row[D_ACT_ARRIVAL]),
            "addr":  short_addr,
        }
        if geocode_source != "por":
            stop["gs"] = geocode_source  # "venue_override" — flag for UI if desired
        route_meta[rn]["stops"].append(stop)
        s_attached += 1

    wb.close()
    print(f"[dispatch] Map_Route_Details: total={s_total:,} attached={s_attached:,} "
          f"skip_no_route={s_no_route:,} skip_no_geo={s_no_geo:,} "
          f"sentinel_coords={s_sentinel:,} "
          f"skip_doc_transfer={s_skip_doc_transfer:,} skip_internal_leg={s_skip_internal:,} "
          f"skip_transfer_leg(cls=T)={s_skip_transfer:,} "
          f"venue_backfill={s_venue_backfill:,} fsa_fallback={s_fsa_fallback:,}", file=sys.stderr)

    # --- Sanity-check start/stop coords against the route's stop bbox ---
    # POR's Map_Route often writes a Burnaby default or stale coord when the
    # depot isn't set, so we only keep start/stop if they sit within ~400 km
    # of the route's stop centroid. 400 km is generous enough to allow for
    # long-haul out-of-city routes without letting a Burnaby coord leak onto
    # a downtown Toronto route.
    THRESH_DEG = 4.0  # ~400 km at Canadian latitudes; cheap lat/lng box check

    def _close_enough(coord, cen_lat, cen_lng):
        return (abs(coord[0] - cen_lat) < THRESH_DEG and
                abs(coord[1] - cen_lng) < THRESH_DEG)

    start_kept = stop_kept = start_rejected = stop_rejected = 0
    start_hs_fallback = stop_hs_fallback = 0
    for rn, meta in route_meta.items():
        stops = meta["stops"]
        if not stops:
            # _start_cand / _stop_cand will be cleaned out when the route drops
            continue
        cen_lat = sum(s["ll"][0] for s in stops) / len(stops)
        cen_lng = sum(s["ll"][1] for s in stops) / len(stops)

        sc = meta.pop("_start_cand", None)
        if sc and _close_enough(sc, cen_lat, cen_lng):
            meta["start"] = [round(sc[0], 6), round(sc[1], 6)]
            start_kept += 1
        elif sc:
            start_rejected += 1

        tc = meta.pop("_stop_cand", None)
        if tc and _close_enough(tc, cen_lat, cen_lng):
            meta["stop"] = [round(tc[0], 6), round(tc[1], 6)]
            stop_kept += 1
        elif tc:
            stop_rejected += 1

        # Fall back to the route's home-store coords whenever POR didn't
        # supply a usable start/stop. Ensures every route's polyline (straight
        # AND OSRM road-following) anchors at the warehouse, not at the first
        # customer stop. The HOME_STORE_COORDS fallback is sanity-checked the
        # same way so we don't paste a Toronto depot on a Halifax route.
        hs_coord = HOME_STORE_COORDS.get(meta.get("hs", ""))
        if hs_coord and _close_enough(hs_coord, cen_lat, cen_lng):
            hs_ll = [round(hs_coord[0], 6), round(hs_coord[1], 6)]
            if "start" not in meta:
                meta["start"] = hs_ll
                start_hs_fallback += 1
            if "stop" not in meta:
                meta["stop"] = hs_ll
                stop_hs_fallback += 1

    print(f"[dispatch] Depot coords: start_kept={start_kept:,} start_rejected={start_rejected:,} "
          f"start_hs_fallback={start_hs_fallback:,} "
          f"stop_kept={stop_kept:,} stop_rejected={stop_rejected:,} "
          f"stop_hs_fallback={stop_hs_fallback:,}", file=sys.stderr)

    # --- Consolidate: drop routes with zero geocoded stops, sort stops by sequence ---
    by_date: dict[str, list] = defaultdict(list)
    dropped_empty = 0
    kept = 0
    for rn, meta in route_meta.items():
        if not meta["stops"]:
            dropped_empty += 1
            continue
        meta["stops"].sort(key=lambda s: s["seq"])
        # Ensure any leftover private keys are removed (safety — they may still
        # be present if a route had 0 stops and skipped the sanity-check loop).
        meta.pop("_start_cand", None)
        meta.pop("_stop_cand", None)
        d = meta.pop("date")
        by_date[d].append(meta)
        kept += 1

    # Sort routes within each day by truck name for consistent rendering
    for d in by_date:
        by_date[d].sort(key=lambda m: (m["truck"] or "", m["rn"]))

    dates_sorted = sorted(by_date.keys())
    print(f"[dispatch] Routes: kept={kept:,} dropped_empty={dropped_empty:,} "
          f"dates={len(dates_sorted)}", file=sys.stderr)

    # --- Optional: fetch road-following geometry via OSRM ---
    # Emits `r.geom` (encoded polyline) per route. Cached to disk so daily runs
    # only hit the network for newly-scheduled routes. Any failure leaves r.geom
    # unset and the client falls back to straight stop-to-stop lines.
    if fetch_geometry:
        # Default: cache next to the output JSON (laptop layout). OSRM_CACHE_PATH
        # overrides for CI, where outputs land at the repo root but the cache is
        # persisted at pipeline/osrm_cache.json between runs via actions/cache.
        cache_path = Path(os.environ.get("OSRM_CACHE_PATH")
                          or Path(out_path).with_name("osrm_cache.json"))
        cache = _osrm_load_cache(cache_path)
        initial_entries = len(cache["entries"])

        route_count = sum(len(v) for v in by_date.values())
        n_hit = n_new = n_fail = n_skip = 0
        seen = 0

        for d in dates_sorted:
            for meta in by_date[d]:
                seen += 1
                # Build waypoint sequence: start depot (if known) → stops in seq → end depot
                wp = []
                if meta.get("start"):
                    wp.append(meta["start"])
                wp.extend([s["ll"] for s in meta["stops"]])
                if meta.get("stop"):
                    wp.append(meta["stop"])

                if len(wp) < 2 or len(wp) > OSRM_MAX_WAYPOINTS:
                    n_skip += 1
                    continue

                geom, was_hit = _fetch_osrm_geometry(wp, osrm_url, cache)
                if geom:
                    meta["geom"] = geom
                if was_hit:
                    n_hit += 1
                elif geom is None:
                    n_fail += 1
                    time.sleep(OSRM_SLEEP_BETWEEN)  # be polite even on failures
                else:
                    n_new += 1
                    time.sleep(OSRM_SLEEP_BETWEEN)

                # Periodic checkpoint — protect against crashes mid-backfill
                if (n_new > 0) and (n_new % 25 == 0) and not was_hit:
                    _osrm_save_cache(cache_path, cache)

                if seen % 50 == 0:
                    print(f"[dispatch] OSRM progress: {seen}/{route_count} "
                          f"(hit={n_hit} new={n_new} fail={n_fail} skip={n_skip})",
                          file=sys.stderr)

        _osrm_save_cache(cache_path, cache)
        added = len(cache["entries"]) - initial_entries
        print(f"[dispatch] OSRM geometry: cache_hits={n_hit:,} new_fetches={n_new:,} "
              f"failed={n_fail:,} skipped={n_skip:,} cache_size={len(cache['entries']):,} "
              f"(+{added:,} new)", file=sys.stderr)

    # --- Contract-truck + event-truck roll-ups ---
    # Lets the UI drill-down show per-truck delivery rate and a multi-truck
    # badge for any contract whose event (date + address) had >1 truck.
    # Keys chosen to keep JSON compact:
    #   contract_trucks[contract] = [[date, truck, route#, cls], ...]
    #   event_trucks[date|addr_key] = list of truck names (unique)
    # The UI groups by (date, addr_key) so linked contracts like L10976 + 691155
    # at 200 University Ave on Apr 15 collapse into a single "2 trucks dispatched"
    # signal rather than being counted separately.
    def _addr_key(addr: str, cust: str) -> str:
        """Normalized key for (address, customer) — loose match that groups
        linked contracts at the same venue on the same day. Prefers the first
        number+street tokens so '200 UNIVERSITY AVE' and '200 UNIVERSITY AVENUE'
        collide. Falls back to customer name when address missing."""
        a = (addr or "").upper().strip()
        if a:
            # Keep only house number + first two street tokens to normalize
            parts = [p for p in a.replace(",", " ").split() if p]
            if parts:
                head = []
                for p in parts:
                    head.append(p)
                    if len(head) >= 3:
                        break
                return " ".join(head)
        return (cust or "").upper().strip()[:40]

    contract_trucks: dict = defaultdict(list)
    event_trucks: dict = defaultdict(set)
    for d in dates_sorted:
        for rmeta in by_date[d]:
            truck = rmeta.get("truck") or ""
            rn = rmeta.get("rn")
            for s in rmeta.get("stops", []):
                con = s.get("con")
                if not con:
                    continue
                contract_trucks[con].append([d, truck, rn, s.get("cls", "D")])
                ek = f"{d}|{_addr_key(s.get('addr',''), s.get('cust',''))}"
                event_trucks[ek].add(truck)

    # Collapse event_trucks to sorted lists for JSON serialization
    event_trucks_out = {k: sorted(v) for k, v in event_trucks.items() if len(v) > 1}

    multi_event_count = len(event_trucks_out)
    multi_contract_count = sum(1 for trucks in contract_trucks.values()
                               if len(set(t[1] for t in trucks)) > 1)
    print(f"[dispatch] Multi-truck: {multi_event_count:,} events have >1 truck "
          f"(grouped by date+addr); {multi_contract_count:,} contracts touched by >1 truck",
          file=sys.stderr)

    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "window": {
            "start": window_start.isoformat(),
            "end":   window_end.isoformat(),
        },
        "dates": dates_sorted,
        "routes": {d: by_date[d] for d in dates_sorted},
        "contract_trucks": dict(contract_trucks),
        "event_trucks":    event_trucks_out,
    }

    # Atomic write: serialize to .tmp, fsync, rename. Prevents partial-file
    # truncation if the process is killed mid-write. Same fix pattern as
    # process_cmmc.py (2026-04-20 hardening).
    out_p = Path(out_path)
    tmp_p = out_p.with_suffix(out_p.suffix + ".tmp")
    data_str = json.dumps(payload, separators=(",", ":"))
    with open(tmp_p, "w", encoding="utf-8") as _fh:
        _fh.write(data_str)
        _fh.flush()
        os.fsync(_fh.fileno())
    os.replace(tmp_p, out_p)
    size_mb = out_p.stat().st_size / 1024 / 1024
    print(f"[dispatch] Wrote {out_path} ({size_mb:.1f} MB)", file=sys.stderr)
    return payload


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("src", help="Path to CMMC_Dispatch_Routes.xlsx")
    ap.add_argument("out", help="Path to dispatch_data.json")
    ap.add_argument("--days-back", type=int, default=90)
    ap.add_argument("--days-forward", type=int, default=30)
    ap.add_argument("--fetch-geometry", action="store_true",
                    help="Fetch road-following geometry via OSRM (cached to osrm_cache.json).")
    ap.add_argument("--osrm-url", default="https://router.project-osrm.org",
                    help="Base URL for OSRM server (default: public demo).")
    args = ap.parse_args()
    process(args.src, args.out, args.days_back, args.days_forward,
            fetch_geometry=args.fetch_geometry, osrm_url=args.osrm_url)


if __name__ == "__main__":
    main()
